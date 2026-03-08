#!/usr/bin/env python3
"""
Import skutečných dat přihlášek 2026 z CERMAT XLSX.

Logika:
1. Přesný match s 2025 ID → plná historie (1484 oborů)
2. Falešná změna (jen diakritika) → propojit na staré ID → plná historie (19 oborů)
3. Nové obory s JPZ, bez historie → jen data 2026, příznak is_new=true (1584 oborů)
4. Obory bez JPZ a bez historie → vyřadit (2902 oborů)

Nikdy se neagreguje přes celou školu!

Aktualizuje:
1. public/applications_2026.json
2. public/school_analysis.json
3. public/schools_data.json
"""

import json
import re
import unicodedata
from pathlib import Path
from collections import defaultdict

import openpyxl

BASE_DIR = Path(__file__).parent.parent
XLSX_PATH = BASE_DIR / 'data' / 'PZ2026_kolo1_skolobory_prihlasky.xlsx'
APP_2026_PATH = BASE_DIR / 'public' / 'applications_2026.json'
SCHOOL_ANALYSIS_PATH = BASE_DIR / 'public' / 'school_analysis.json'
SCHOOLS_DATA_PATH = BASE_DIR / 'public' / 'schools_data.json'


def slugify(text: str) -> str:
    """Převede text na slug."""
    if not text:
        return ''
    text = unicodedata.normalize('NFKD', text)
    text = text.encode('ascii', 'ignore').decode('ascii')
    text = re.sub(r'[^a-zA-Z0-9]+', '_', text)
    text = text.strip('_')
    return text


def normalize_slug(slug: str) -> str:
    """Normalizuje slug pro porovnání (lowercase, bez oddělovačů)."""
    return slug.lower().replace('_', '').replace('-', '')


def make_full_id(redizo, kkov, zamereni):
    """Vytvoří full ID z REDIZO, KKOV a zaměření."""
    base = f"{redizo}_{kkov}"
    zam_slug = slugify(zamereni)
    return f"{base}_{zam_slug}" if zam_slug else base


def zrizovatel_label(code):
    """Mapuje kód zřizovatele na label."""
    mapping = {
        '1': 'veřejné / státní', 'MŠMT': 'veřejné / státní',
        '2': 'veřejné / státní', 'krajské': 'veřejné / státní',
        '3': 'veřejné / státní', 'obecní': 'veřejné / státní',
        '5': 'soukromé', 'soukromé': 'soukromé',
        '6': 'církevní', 'církevní': 'církevní',
    }
    return mapping.get(str(code), str(code) if code else 'veřejné / státní')


def load_cermat_xlsx():
    """Načte CERMAT XLSX a vrátí filtrované záznamy (denní, nezkrácené)."""
    print(f"Načítám {XLSX_PATH}...")
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    print(f"  Celkem řádků: {len(rows) - 1}")

    records = []
    for r in rows[1:]:
        forma = r[27]   # FORMA VZDĚLÁVÁNÍ
        zkracene = r[29] # ZKRÁCENÉ STUDIUM
        if forma == 'den' and zkracene == 'ne':
            records.append({
                'redizo': str(r[5]),
                'nazev': str(r[6]) if r[6] else '',
                'ulice': str(r[7]) if r[7] else '',
                'obec': str(r[8]) if r[8] else '',
                'psc': str(r[9]) if r[9] else '',
                'kraj_kod': str(r[10]) if r[10] else '',
                'kraj': str(r[11]) if r[11] else '',
                'okres': str(r[13]) if r[13] else '',
                'orp': str(r[15]) if r[15] else '',
                'zrizovatel': str(r[17]) if r[17] else '',
                'jpz': int(r[20]) if r[20] else 0,
                'typ': str(r[21]) if r[21] else '',
                'kkov': str(r[24]) if r[24] else '',
                'obor': str(r[25]) if r[25] else '',
                'zamereni': str(r[26]) if r[26] else '',
                'delka_studia': int(r[28]) if r[28] else 4,
                'kapacita': int(r[31]) if r[31] else 0,
                'prihlasky': int(r[33]) if r[33] else 0,
                'p1': int(r[34]) if r[34] else 0,
                'p2': int(r[35]) if r[35] else 0,
                'p3': int(r[36]) if r[36] else 0,
                'p4': int(r[37]) if r[37] else 0,
                'p5': int(r[38]) if r[38] else 0,
            })

    print(f"  Denní nezkrácené: {len(records)}")
    return records


def load_orig_2025():
    """Načte původní 2025 data (ty s reálnými čísly)."""
    with open(SCHOOLS_DATA_PATH, 'r', encoding='utf-8') as f:
        sd = json.load(f)
    # Původní záznamy = ty co mají data
    orig = {r['id']: r for r in sd['2025']
            if r.get('prihlasky', 0) > 0 or r.get('kapacita', 0) > 0}
    print(f"  Původní 2025 záznamů s daty: {len(orig)}")
    return orig


def classify_records(records, orig_2025):
    """Klasifikuje CERMAT záznamy do kategorií."""
    # Build normalized index: base -> {normalized_slug: orig_id}
    orig_base_norm = {}
    for fid in orig_2025:
        parts = fid.split('_')
        base = parts[0] + '_' + parts[1]
        zam = '_'.join(parts[2:]) if len(parts) > 2 else ''
        norm = normalize_slug(zam)
        if base not in orig_base_norm:
            orig_base_norm[base] = {}
        orig_base_norm[base][norm] = fid

    exact_match = []      # Přesný match → plná historie
    false_change = []     # Falešná změna (diakritika) → propojit
    new_with_jpz = []     # Nové s JPZ
    skipped_no_jpz = []   # Bez JPZ, bez historie → vyřadit

    for r in records:
        cermat_id = make_full_id(r['redizo'], r['kkov'], r['zamereni'])
        base = f"{r['redizo']}_{r['kkov']}"
        zam_slug = slugify(r['zamereni'])
        norm = normalize_slug(zam_slug)

        if cermat_id in orig_2025:
            # Přesný match
            r['_final_id'] = cermat_id
            r['_has_history'] = True
            r['_is_new'] = False
            exact_match.append(r)
        else:
            # Zkusit normalized match (falešná změna)
            base_norms = orig_base_norm.get(base, {})
            if norm in base_norms:
                # Falešná změna - použít STARÉ ID pro propojení historie
                old_id = base_norms[norm]
                r['_final_id'] = old_id
                r['_has_history'] = True
                r['_is_new'] = False
                r['_false_change_from'] = cermat_id
                false_change.append(r)
            elif r['jpz'] == 1:
                # Nový obor s JPZ
                r['_final_id'] = cermat_id
                r['_has_history'] = False
                r['_is_new'] = True
                new_with_jpz.append(r)
            else:
                # Bez JPZ, bez historie → vyřadit
                skipped_no_jpz.append(r)

    print(f"\n=== KLASIFIKACE ===")
    print(f"  Přesný match (plná historie):    {len(exact_match):>5}")
    print(f"  Falešná změna (propojeno):       {len(false_change):>5}")
    print(f"  Nové s JPZ (jen 2026 data):      {len(new_with_jpz):>5}")
    print(f"  Vyřazeno (bez JPZ, bez hist.):   {len(skipped_no_jpz):>5}")
    print(f"  ─────────────────────────────────────")
    print(f"  DO SYSTÉMU:                      {len(exact_match) + len(false_change) + len(new_with_jpz):>5}")

    if false_change:
        print(f"\n  Falešné změny (ukázky):")
        for r in false_change[:5]:
            print(f"    {r['_false_change_from']} → {r['_final_id']}")

    return exact_match + false_change + new_with_jpz


def build_app_2026_records(classified):
    """Vytvoří záznamy pro applications_2026.json."""
    # Deduplicate by final_id (sčítat pokud se potkají)
    merged = {}
    for r in classified:
        fid = r['_final_id']
        if fid in merged:
            m = merged[fid]
            m['kapacita'] += r['kapacita']
            m['prihlasky'] += r['prihlasky']
            m['pp'][0] += r['p1']
            m['pp'][1] += r['p2']
            m['pp'][2] += r['p3']
            m['pp'][3] += r['p4']
            m['pp'][4] += r['p5']
        else:
            merged[fid] = {
                'id': fid,
                'kapacita': r['kapacita'],
                'prihlasky': r['prihlasky'],
                'pp': [r['p1'], r['p2'], r['p3'], r['p4'], r['p5']],
                'is_new': r['_is_new'],
                # Statická data pro nové obory
                '_raw': r,
            }

    # Dopočítat idx
    for m in merged.values():
        m['idx'] = round(m['prihlasky'] / m['kapacita'], 2) if m['kapacita'] > 0 else 0

    return merged


def write_applications_2026(merged):
    """Zapíše applications_2026.json."""
    print(f"\nZapisuji {APP_2026_PATH}...")

    records = []
    for m in merged.values():
        rec = {
            'id': m['id'],
            'kapacita': m['kapacita'],
            'prihlasky': m['prihlasky'],
            'pp': m['pp'],
            'idx': m['idx'],
        }
        if m['is_new']:
            rec['is_new'] = True
        records.append(rec)

    total_prihl = sum(r['prihlasky'] for r in records)
    new_count = sum(1 for r in records if r.get('is_new'))

    output = {
        'meta': {
            'rok': 2026,
            'kolo': 1,
            'celkem_prihlasek': total_prihl,
            'celkem_oboru': len(records),
            'zdroj': 'CERMAT data.cermat.cz, aktualizace 2026-03-08',
        },
        'data': records,
    }

    with open(APP_2026_PATH, 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, separators=(',', ':'))

    size_kb = APP_2026_PATH.stat().st_size / 1024
    print(f"  Zapsáno {len(records)} záznamů ({size_kb:.0f} KB)")
    print(f"  S historií: {len(records) - new_count}, nové: {new_count}")
    print(f"  Celkem přihlášek: {total_prihl:,}")


def write_school_analysis(merged):
    """Aktualizuje school_analysis.json."""
    print(f"\nAktualizuji {SCHOOL_ANALYSIS_PATH}...")

    # Načíst ČISTÝ soubor - jen původní 2025 data
    # Nejprve vrátíme school_analysis do původního stavu (bez _2026 polí)
    with open(SCHOOL_ANALYSIS_PATH, 'r', encoding='utf-8') as f:
        sa = json.load(f)

    # Odstranit předchozí importované záznamy a pole
    keys_to_remove = []
    for key in sa:
        # Odstranit _2026 pole z existujících
        for field in ['prihlasky_2026', 'kapacita_2026', 'index_poptavky_2026', 'prihlasky_priority_2026']:
            sa[key].pop(field, None)
        # Označit záznamy přidané předchozím importem
        if sa[key].get('category_name') == 'Nový obor':
            keys_to_remove.append(key)
    for key in keys_to_remove:
        del sa[key]

    # Agregovat merged záznamy podle base klíče (REDIZO_KKOV) pro school_analysis
    base_agg = defaultdict(lambda: {
        'kapacita': 0, 'prihlasky': 0,
        'pp': [0, 0, 0, 0, 0], 'is_new': True,
    })
    base_raw = {}  # Uložit raw data pro nové obory

    for m in merged.values():
        parts = m['id'].split('_')
        base = parts[0] + '_' + parts[1]
        agg = base_agg[base]
        agg['kapacita'] += m['kapacita']
        agg['prihlasky'] += m['prihlasky']
        for i in range(5):
            agg['pp'][i] += m['pp'][i]
        if not m['is_new']:
            agg['is_new'] = False
        if base not in base_raw:
            base_raw[base] = m['_raw']

    matched = 0
    new_count = 0

    for base, agg in base_agg.items():
        idx = round(agg['prihlasky'] / agg['kapacita'], 2) if agg['kapacita'] > 0 else 0

        if base in sa:
            sa[base]['prihlasky_2026'] = agg['prihlasky']
            sa[base]['kapacita_2026'] = agg['kapacita']
            sa[base]['index_poptavky_2026'] = idx
            sa[base]['prihlasky_priority_2026'] = agg['pp']
            if agg['is_new']:
                sa[base]['is_new_2026'] = True
            matched += 1
        else:
            # Nový obor
            raw = base_raw.get(base, {})
            sa[base] = {
                'id': base,
                'nazev': raw.get('nazev', ''),
                'obor': raw.get('obor', ''),
                'obec': raw.get('obec', ''),
                'okres': raw.get('okres', ''),
                'orp': raw.get('orp', ''),
                'kraj': raw.get('kraj', ''),
                'kraj_kod': raw.get('kraj_kod', ''),
                'adresa': f"{raw.get('ulice', '')}, {raw.get('obec', '')}, {raw.get('psc', '')}",
                'adresa_plna': f"{raw.get('ulice', '')}, {raw.get('obec', '')}, {raw.get('psc', '')}",
                'zrizovatel': zrizovatel_label(raw.get('zrizovatel', '')),
                'typ': raw.get('typ', ''),
                'delka_studia': raw.get('delka_studia', 4),
                'min_body': 0, 'prumer_body': 0,
                'kapacita': 0, 'prihlasky': 0, 'prijati': 0,
                'index_poptavky': 0, 'obtiznost': 0,
                'total_applicants': 0,
                'priority_counts': [0, 0, 0, 0, 0],
                'priority_pcts': [0, 0, 0, 0, 0],
                'category_code': 'balanced',
                'category_name': 'Nový obor 2026',
                'prihlasky_2026': agg['prihlasky'],
                'kapacita_2026': agg['kapacita'],
                'index_poptavky_2026': idx,
                'prihlasky_priority_2026': agg['pp'],
                'is_new_2026': True,
            }
            new_count += 1

    with open(SCHOOL_ANALYSIS_PATH, 'w', encoding='utf-8') as f:
        json.dump(sa, f, ensure_ascii=False, separators=(',', ':'))

    size_mb = SCHOOL_ANALYSIS_PATH.stat().st_size / 1024 / 1024
    print(f"  Aktualizováno: {matched}, nových: {new_count}")
    print(f"  Celkem: {len(sa)} ({size_mb:.1f} MB)")


def write_schools_data(merged, orig_2025):
    """Aktualizuje schools_data.json - přidá nové obory."""
    print(f"\nAktualizuji {SCHOOLS_DATA_PATH}...")

    with open(SCHOOLS_DATA_PATH, 'r', encoding='utf-8') as f:
        sd = json.load(f)

    # Vyčistit: ponechat jen původní záznamy (ty s daty)
    sd['2025'] = [r for r in sd['2025']
                  if r.get('prihlasky', 0) > 0 or r.get('kapacita', 0) > 0]
    existing_ids = {r['id'] for r in sd['2025']}

    new_count = 0
    for m in merged.values():
        fid = m['id']
        if fid not in existing_ids and m['is_new']:
            raw = m['_raw']
            sd['2025'].append({
                'id': fid,
                'redizo': raw['redizo'],
                'nazev': raw['nazev'],
                'nazev_display': f"{raw['nazev']}, {raw['ulice']}" if raw.get('ulice') else raw['nazev'],
                'adresa_plna': f"{raw.get('ulice', '')}, {raw['obec']}, {raw.get('psc', '')}",
                'ulice': raw.get('ulice', ''),
                'obec': raw['obec'],
                'psc': raw.get('psc', ''),
                'kraj_kod': raw['kraj_kod'],
                'kraj': raw['kraj'],
                'okres': raw.get('okres', ''),
                'orp': raw.get('orp', ''),
                'zrizovatel': zrizovatel_label(raw.get('zrizovatel', '')),
                'obor': raw['obor'],
                'zamereni': raw['zamereni'],
                'kkov': raw['kkov'],
                'typ': raw['typ'],
                'delka_studia': raw['delka_studia'],
                'adresa': f"{raw.get('ulice', '')}, {raw['obec']}, {raw.get('psc', '')}",
                'kapacita': 0, 'prihlasky': 0, 'prijati': 0,
                'index_poptavky': 0, 'min_body': 0, 'prumer_body': 0,
                'rok': 2025,
                'is_new_2026': True,
            })
            new_count += 1

    with open(SCHOOLS_DATA_PATH, 'w', encoding='utf-8') as f:
        json.dump(sd, f, ensure_ascii=False, separators=(',', ':'))

    size_mb = SCHOOLS_DATA_PATH.stat().st_size / 1024 / 1024
    print(f"  Původní 2025: {len(existing_ids)}, nových: {new_count}")
    print(f"  Celkem 2025: {len(sd['2025'])} ({size_mb:.1f} MB)")


def verify_issues(merged):
    """Ověří data pro nahlášené issues."""
    print("\n=== OVĚŘENÍ ISSUES ===")

    # Agregovat per base pro ověření
    base_prihl = defaultdict(int)
    base_kap = defaultdict(int)
    for m in merged.values():
        parts = m['id'].split('_')
        base = parts[0] + '_' + parts[1]
        base_prihl[base] += m['prihlasky']
        base_kap[base] += m['kapacita']

    checks = [
        ('691000468_79-41-K/41', '#43 Gymnázium M. Horákové', 188),
        ('600005933_79-41-K/41', '#45 Gymnázium U lib. zámku', 427),
    ]
    for base, label, expected in checks:
        p = base_prihl.get(base, '?')
        k = base_kap.get(base, '?')
        status = '✅' if p == expected else '❌'
        print(f"  {status} {label}: prihl={p}, kap={k} (očekáváno {expected})")

    # #44 Drtinova
    for base in sorted(base_prihl):
        if base.startswith('600005674'):
            print(f"  #44 Drtinova: {base} prihl={base_prihl[base]}, kap={base_kap[base]}")


def print_examples(merged):
    """Vypíše příklady pro každou kategorii."""
    print("\n=== VZOROVÉ OBORY ===")
    with_hist = [(m['id'], m) for m in merged.values() if not m['is_new']]
    new_obory = [(m['id'], m) for m in merged.values() if m['is_new']]

    print(f"\nS historií ({len(with_hist)}):")
    for fid, m in sorted(with_hist, key=lambda x: -x[1]['prihlasky'])[:3]:
        r = m['_raw']
        print(f"  {fid}: {r['nazev']} / {r['obor']} / {r.get('zamereni','')} "
              f"| prihl={m['prihlasky']}, kap={m['kapacita']}")

    print(f"\nNové s JPZ ({len(new_obory)}):")
    for fid, m in sorted(new_obory, key=lambda x: -x[1]['prihlasky'])[:3]:
        r = m['_raw']
        print(f"  {fid}: {r['nazev']} / {r['obor']} / {r.get('zamereni','')} "
              f"| prihl={m['prihlasky']}, kap={m['kapacita']}")


def main():
    print("=" * 60)
    print("Import skutečných dat CERMAT 2026 (v2)")
    print("=" * 60)

    # 1. Načíst data
    records = load_cermat_xlsx()
    print("\nNačítám původní 2025 data...")
    orig_2025 = load_orig_2025()

    # 2. Klasifikovat
    classified = classify_records(records, orig_2025)

    # 3. Vytvořit merged záznamy
    merged = build_app_2026_records(classified)
    print(f"\nMerged záznamů (po deduplikaci): {len(merged)}")

    # 4. Ověřit issues
    verify_issues(merged)

    # 5. Příklady
    print_examples(merged)

    # 6. Zapsat soubory
    write_applications_2026(merged)
    write_school_analysis(merged)
    write_schools_data(merged, orig_2025)

    print("\n" + "=" * 60)
    print("HOTOVO!")
    print("=" * 60)


if __name__ == '__main__':
    main()
