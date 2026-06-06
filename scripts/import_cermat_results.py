#!/usr/bin/env python3
"""
Import výsledků přijímacích zkoušek z CERMAT XLSX.

Použití:
    python scripts/import_cermat_results.py --year 2026

Vstup:  cermat_data_2025/PZ{YEAR}_kolo1_vysledky.xlsx
        cermat_data_2025/PZ{YEAR-1}_kolo1_vysledky.xlsx  (pro Δ hodnoty)
Výstup: public/cermat_results_{YEAR}.json
        public/cermat_results_meta.json
"""
import argparse
import json
import re
import unicodedata
from pathlib import Path
from collections import defaultdict

import openpyxl

BASE_DIR = Path(__file__).parent.parent
CERMAT_DIR = BASE_DIR / 'cermat_data_2025'
PUBLIC_DIR = BASE_DIR / 'public'
META_PATH = PUBLIC_DIR / 'cermat_results_meta.json'


def slugify(text: str) -> str:
    if not text:
        return ''
    text = unicodedata.normalize('NFKD', str(text))
    text = text.encode('ascii', 'ignore').decode('ascii')
    text = re.sub(r'[^a-zA-Z0-9]+', '_', text)
    return text.strip('_').lower()


def make_key(redizo, kkov, zamereni=''):
    base = f"{redizo}_{kkov}"
    z = slugify(zamereni)
    return f"{base}_{z}" if z else base


def load_flat_xlsx(path: Path) -> list[dict]:
    """Načte flat xlsx (2026 formát) — 1 list, hlavičky na řádku 1."""
    print(f"Načítám {path.name}...")
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h) if h else '' for h in rows[0]]
    records = []
    for row in rows[1:]:
        if row[0] is None:
            continue
        records.append(dict(zip(headers, row)))
    print(f"  Načteno {len(records)} řádků")
    return records


def load_regional_xlsx(path: Path) -> list[dict]:
    """Načte xlsx s regionálními listy (2025 formát) — hlavičky na řádku 2."""
    SHEETS = [
        'CZ010-PHA', 'CZ020-SCK', 'CZ031-JCK', 'CZ032-PLK', 'CZ041-KVK',
        'CZ042-USK', 'CZ051-LBK', 'CZ052-KHK', 'CZ053-PAK', 'CZ063-VYS',
        'CZ064-JMK', 'CZ071-OLK', 'CZ072-ZLK', 'CZ080-MSK'
    ]
    print(f"Načítám {path.name} (regionální listy)...")
    wb = openpyxl.load_workbook(path, read_only=True)
    records = []
    for sn in SHEETS:
        if sn not in wb.sheetnames:
            continue
        ws = wb[sn]
        headers = None
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 1:
                headers = [str(h) if h else '' for h in row]
            elif i >= 2 and headers and row[0] is not None:
                records.append(dict(zip(headers, row)))
    print(f"  Načteno {len(records)} řádků ze {len(SHEETS)} listů")
    return records


def is_valid_flat(r: dict) -> bool:
    """Filtr pro 2026 (flat) záznamy: denní, nezkrácené, s JPZ."""
    forma = str(r.get('FORMA VZDĚLÁVÁNÍ') or '').lower()
    zkracene = str(r.get('ZKRÁCENÉ STUDIUM') or '').lower()
    jpz = r.get('POVINNOST JPZ')
    return 'den' in forma and zkracene == 'ne' and jpz == 1


def is_valid_regional(r: dict) -> bool:
    """Filtr pro 2025 (regional) záznamy: denní, nezkrácené."""
    forma = str(r.get('FORMA') or '').lower()
    zkracene = str(r.get('ZKRÁCENÉ STUDIUM') or '').lower()
    return 'den' in forma and zkracene in ('', 'ne', 'false', '0')


def safe_float(v, default=0.0) -> float:
    if v is None or isinstance(v, str):
        return default
    try:
        return float(v)
    except (TypeError, ValueError):
        return default


def extract_current_year(records: list[dict]) -> dict[str, dict]:
    """Vrátí dict klíčovaný make_key → výsledky z flat xlsx (aktuální rok)."""
    result = {}
    for r in records:
        if not is_valid_flat(r):
            continue
        redizo = str(r.get('REDIZO') or '')
        kkov = str(r.get('KKOV') or '')
        zamereni = str(r.get('ZAMĚŘENÍ OBORU') or '')
        if not redizo or not kkov:
            continue
        key = make_key(redizo, kkov, zamereni)
        # CERMAT dodává % skór (0-100 na předmět, 0-200 celkem).
        # Pro konzistenci se zbytkem aplikace převádíme na body z testu (0-50 / 0-100).
        cj_ma_pct = safe_float(r.get('ČJ+MA - % SKÓR - PRŮMĚR (PŘIJATI)'))
        cj_pct = safe_float(r.get('ČJ - % SKÓR - PRŮMĚR (PŘIJATI)'))
        ma_pct = safe_float(r.get('MA - % SKÓR - PRŮMĚR (PŘIJATI)'))
        if cj_ma_pct == 0:
            continue
        result[key] = {
            'redizo': redizo,
            'kkov': kkov,
            'zamereni': zamereni,
            'nazev': str(r.get('NÁZEV ŠKOLY') or ''),
            'kraj': str(r.get('KRAJ - NÁZEV') or ''),
            'school_type': str(r.get('TYP ŠKOLY') or ''),
            'kapacita': int(safe_float(r.get('KAPACITA'))),
            'prijati': int(safe_float(r.get('PŘIJATÍ'))),
            'cj_ma_prijati': round(cj_ma_pct / 2, 2),
            'cj_prijati': round(cj_pct / 2, 2),
            'ma_prijati': round(ma_pct / 2, 2),
        }
    return result


def extract_prev_year(records: list[dict]) -> dict[str, float]:
    """Vrátí dict klíčovaný make_key → cj_ma_prijati z regionálního xlsx (předchozí rok)."""
    result = {}
    for r in records:
        if not is_valid_regional(r):
            continue
        redizo = str(r.get('REDIZO') or '')
        kkov = str(r.get('KKOV') or '')
        zamereni = str(r.get('ZAMĚŘENÍ') or '')
        if not redizo or not kkov:
            continue
        key = make_key(redizo, kkov, zamereni)
        # Stejný převod jako u flat: % skór (0-200) → body (0-100)
        v_pct = safe_float(r.get('ČJ+MA - % skór - průměr - přijati'))
        if v_pct > 0:
            result[key] = round(v_pct / 2, 2)
    return result


def compute_ranks(data: dict[str, dict]) -> dict[str, dict]:
    """Přidá rank_in_type a type_total ke každému záznamu."""
    by_type: dict[str, list] = defaultdict(list)
    for key, rec in data.items():
        by_type[rec['school_type']].append((key, rec['cj_ma_prijati']))
    for typ, items in by_type.items():
        items.sort(key=lambda x: -x[1])
        for rank, (key, _) in enumerate(items, 1):
            data[key]['rank_in_type'] = rank
            data[key]['type_total'] = len(items)
    return data


def enrich_nazev_display(data: dict[str, dict]) -> dict[str, dict]:
    """Doplní nazev_display z schools_data.json (nazev + ulice pro disambiguaci)."""
    sd_path = PUBLIC_DIR / 'schools_data.json'
    if not sd_path.exists():
        print("schools_data.json nenalezen — nazev_display nebude obohacen")
        return data

    sd = json.loads(sd_path.read_text())
    lookup: dict[str, str] = {}
    for year_data in sd.values():
        if not isinstance(year_data, list):
            continue
        for school in year_data:
            redizo = school.get('redizo', '')
            nd = school.get('nazev_display', '')
            if redizo and nd:
                lookup[redizo] = nd

    enriched = 0
    for rec in data.values():
        if rec.get('nazev_display'):
            continue
        nd = lookup.get(rec['redizo'], '')
        if nd:
            rec['nazev_display'] = nd
            enriched += 1
        else:
            rec['nazev_display'] = rec['nazev']
    print(f"Obohaceno nazev_display: {enriched}/{len(data)}")
    return data


def update_meta(year: int, meta_path: Path) -> None:
    if meta_path.exists():
        meta = json.loads(meta_path.read_text())
    else:
        meta = {'latest_year': year, 'available_years': []}
    if year not in meta['available_years']:
        meta['available_years'].append(year)
        meta['available_years'].sort()
    meta['latest_year'] = max(meta['available_years'])
    meta_path.write_text(json.dumps(meta, ensure_ascii=False, indent=2))
    print(f"Meta aktualizováno: {meta}")


def main(year: int) -> None:
    curr_path = CERMAT_DIR / f'PZ{year}_kolo1_vysledky.xlsx'
    if not curr_path.exists():
        raise FileNotFoundError(f"Soubor nenalezen: {curr_path}")
    curr_records = load_flat_xlsx(curr_path)
    data = extract_current_year(curr_records)
    print(f"Extrahováno {len(data)} platných záznamů z {year}")

    prev_year = year - 1
    prev_path = CERMAT_DIR / f'PZ{prev_year}_kolo1_vysledky.xlsx'
    prev_data: dict[str, float] = {}
    if prev_path.exists():
        wb_prev = openpyxl.load_workbook(prev_path, read_only=True, data_only=True)
        sheets = wb_prev.sheetnames
        wb_prev.close()
        if len(sheets) == 1:
            prev_records = load_flat_xlsx(prev_path)
            prev_flat = extract_current_year(prev_records)
            prev_data = {k: v['cj_ma_prijati'] for k, v in prev_flat.items()}
        else:
            prev_records = load_regional_xlsx(prev_path)
            prev_data = extract_prev_year(prev_records)
        print(f"Načteno {len(prev_data)} prev záznamů z {prev_year}")
    else:
        print(f"Soubor {prev_path.name} nenalezen — Δ hodnoty nebudou k dispozici")

    matched = 0
    for key, rec in data.items():
        prev_val = prev_data.get(key, 0.0)
        rec['cj_ma_prijati_prev'] = prev_val
        rec['delta_cj_ma'] = round(rec['cj_ma_prijati'] - prev_val, 2) if prev_val > 0 else None
        if prev_val > 0:
            matched += 1
    print(f"Spárováno s předchozím rokem: {matched}/{len(data)}")

    data = compute_ranks(data)
    data = enrich_nazev_display(data)

    out_path = PUBLIC_DIR / f'cermat_results_{year}.json'
    out_path.write_text(json.dumps(data, ensure_ascii=False, indent=2))
    print(f"Zapsáno: {out_path} ({len(data)} záznamů)")

    update_meta(year, META_PATH)


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--year', type=int, required=True)
    args = parser.parse_args()
    main(args.year)
